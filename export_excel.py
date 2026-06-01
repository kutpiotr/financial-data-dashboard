"""
export_excel.py
---------------
Generuje plik financial_report.xlsx z trzema arkuszami:
1. Summary NAV        — ostatnie ceny ETF-ów, zmiany dzienne i YTD
2. Exchange Rates     — tabela kursów walut z odchyleniami od średniej
3. Monthly Returns    — pivot z miesięcznymi zwrotami procentowymi

Formatowanie warunkowe:
- Zielone/czerwone tło dla dodatnich/ujemnych zwrotów
- Pomarańczowe tło dla odchyleń kursów walut > 2%
"""

import os
import logging
import sqlite3
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, "export_excel.log")),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

DB_PATH = "finance.db"
OUTPUT_FILE = "financial_report.xlsx"

# Kolory
COLOR_HEADER_BG  = "1F4E79"   # ciemny niebieski
COLOR_HEADER_FG  = "FFFFFF"   # biały
COLOR_POS        = "C6EFCE"   # jasny zielony
COLOR_NEG        = "FFC7CE"   # jasny czerwony
COLOR_ALERT      = "FFEB9C"   # pomarańczowy/żółty
COLOR_ALT_ROW    = "EBF3FB"   # jasny niebieski (co drugi wiersz)
COLOR_BORDER     = "B8CCE4"


# ── Helpers ──────────────────────────────────────────────────────────────────

def thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)


def header_font() -> Font:
    return Font(bold=True, color=COLOR_HEADER_FG, size=10)


def alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_ALT_ROW)


def style_header_row(ws, row_idx: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = header_fill()
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()


def autofit_columns(ws, min_width: int = 10, max_width: int = 30) -> None:
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            min_width, min(length + 2, max_width)
        )


# ── Dane z bazy ──────────────────────────────────────────────────────────────

def read_sql(query: str, db: str = DB_PATH) -> pd.DataFrame:
    if not os.path.exists(db):
        logger.warning(f"Baza nie istnieje: {db} — używam pustych danych demo")
        return pd.DataFrame()
    conn = sqlite3.connect(db)
    try:
        return pd.read_sql_query(query, conn)
    finally:
        conn.close()


def get_nav_data() -> pd.DataFrame:
    sql = """
    WITH latest AS (
        SELECT ticker, close AS latest_close, date AS latest_date
        FROM prices
        WHERE (ticker, date) IN (SELECT ticker, MAX(date) FROM prices GROUP BY ticker)
    ),
    start_of_year AS (
        SELECT ticker, close AS ytd_open
        FROM prices
        WHERE (ticker, date) IN (
            SELECT ticker, MIN(date) FROM prices
            WHERE STRFTIME('%Y', date) = STRFTIME('%Y', 'now')
            GROUP BY ticker
        )
    ),
    prev_day AS (
        SELECT p1.ticker, p1.close AS prev_close
        FROM prices p1
        WHERE p1.date = (
            SELECT MAX(p2.date) FROM prices p2
            WHERE p2.ticker = p1.ticker
              AND p2.date < (SELECT MAX(date) FROM prices WHERE ticker = p1.ticker)
        )
    )
    SELECT
        l.ticker, l.latest_date AS date,
        ROUND(l.latest_close, 4) AS price,
        ROUND(p.prev_close, 4) AS prev_price,
        ROUND(l.latest_close - p.prev_close, 4) AS day_change,
        ROUND((l.latest_close - p.prev_close) / p.prev_close * 100, 2) AS day_change_pct,
        ROUND(s.ytd_open, 4) AS ytd_open,
        ROUND((l.latest_close - s.ytd_open) / s.ytd_open * 100, 2) AS ytd_return_pct
    FROM latest l
    LEFT JOIN start_of_year s ON l.ticker = s.ticker
    LEFT JOIN prev_day p ON l.ticker = p.ticker
    ORDER BY l.ticker
    """
    df = read_sql(sql)
    if df.empty:
        # Dane demo
        df = pd.DataFrame({
            "ticker": ["CSPX.L", "SPY", "VWRA.L"],
            "date": ["2024-01-15"] * 3,
            "price": [450.20, 480.50, 112.30],
            "prev_price": [448.10, 477.20, 111.80],
            "day_change": [2.10, 3.30, 0.50],
            "day_change_pct": [0.47, 0.69, 0.45],
            "ytd_open": [430.00, 460.00, 108.00],
            "ytd_return_pct": [4.70, 4.46, 3.98],
        })
    return df


def get_fx_data() -> pd.DataFrame:
    sql = """
    SELECT date, pair, close,
        ROUND(AVG(close) OVER (PARTITION BY pair ORDER BY date ROWS BETWEEN 29 PRECEDING AND CURRENT ROW), 4) AS avg_30d,
        ROUND((close - AVG(close) OVER (PARTITION BY pair ORDER BY date ROWS BETWEEN 29 PRECEDING AND CURRENT ROW))
              / AVG(close) OVER (PARTITION BY pair ORDER BY date ROWS BETWEEN 29 PRECEDING AND CURRENT ROW) * 100, 2) AS pct_dev
    FROM exchange_rates
    ORDER BY pair, date DESC
    LIMIT 200
    """
    df = read_sql(sql)
    if df.empty:
        import numpy as np
        dates = pd.date_range("2024-01-01", periods=60, freq="B")
        rows = []
        for pair, base in [("USD_PLN", 4.0), ("EUR_PLN", 4.3)]:
            prices = base + np.random.randn(60).cumsum() * 0.01
            for d, p in zip(dates, prices):
                rows.append({"date": str(d.date()), "pair": pair, "close": round(p, 4),
                             "avg_30d": round(p * 0.998, 4), "pct_dev": round((p / (p * 0.998) - 1) * 100, 2)})
        df = pd.DataFrame(rows)
    return df


def get_monthly_returns() -> pd.DataFrame:
    sql = """
    WITH monthly AS (
        SELECT ticker AS symbol,
               STRFTIME('%Y-%m', date) AS year_month,
               FIRST_VALUE(close) OVER (PARTITION BY ticker, STRFTIME('%Y-%m', date) ORDER BY date) AS open_p,
               LAST_VALUE(close)  OVER (PARTITION BY ticker, STRFTIME('%Y-%m', date) ORDER BY date
                                        ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING) AS close_p,
               ROW_NUMBER() OVER (PARTITION BY ticker, STRFTIME('%Y-%m', date) ORDER BY date DESC) AS rn
        FROM prices
        UNION ALL
        SELECT symbol,
               STRFTIME('%Y-%m', date),
               FIRST_VALUE(close) OVER (PARTITION BY symbol, STRFTIME('%Y-%m', date) ORDER BY date),
               LAST_VALUE(close)  OVER (PARTITION BY symbol, STRFTIME('%Y-%m', date) ORDER BY date
                                        ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING),
               ROW_NUMBER() OVER (PARTITION BY symbol, STRFTIME('%Y-%m', date) ORDER BY date DESC)
        FROM indices
    )
    SELECT symbol, year_month,
           ROUND((close_p - open_p) / open_p * 100, 2) AS monthly_return_pct
    FROM monthly WHERE rn = 1
    ORDER BY symbol, year_month
    """
    df = read_sql(sql)
    if df.empty:
        import numpy as np
        months = pd.period_range("2023-01", periods=24, freq="M").astype(str)
        rows = []
        for sym in ["CSPX.L", "SPY", "VWRA.L", "^GSPC"]:
            for m in months:
                rows.append({"symbol": sym, "year_month": m,
                             "monthly_return_pct": round(float(np.random.randn() * 3), 2)})
        df = pd.DataFrame(rows)
    return df


# ── Arkusze ───────────────────────────────────────────────────────────────────

def write_nav_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Summary NAV")
    ws.freeze_panes = "A2"

    col_names = ["Ticker", "Data", "Cena", "Poprzednia cena",
                 "Zmiana dzienna", "Zmiana % (dzienna)",
                 "Cena pocz. roku", "Zwrot YTD (%)"]
    df.columns = col_names[:len(df.columns)]

    # Nagłówek
    ws.append(col_names)
    style_header_row(ws, 1, len(col_names))

    # Dane
    pos_fill = PatternFill("solid", fgColor=COLOR_POS)
    neg_fill = PatternFill("solid", fgColor=COLOR_NEG)

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        for col in range(1, len(col_names) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = alt_fill()

        # Kolorowanie zmiany dziennej i YTD
        for col_idx in (6, 8):  # Zmiana % (dzienna), Zwrot YTD
            val = ws.cell(row=i, column=col_idx).value
            if val is not None:
                ws.cell(row=i, column=col_idx).fill = pos_fill if val >= 0 else neg_fill
                ws.cell(row=i, column=col_idx).number_format = '0.00"%"'

    autofit_columns(ws)
    ws.column_dimensions["A"].width = 12


def write_fx_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Exchange Rates")
    ws.freeze_panes = "A2"

    col_names = ["Data", "Para walutowa", "Kurs", "Śr. 30-dniowa", "Odchylenie %"]
    df.columns = col_names[:len(df.columns)]

    ws.append(col_names)
    style_header_row(ws, 1, len(col_names))

    alert_fill = PatternFill("solid", fgColor=COLOR_ALERT)
    pos_fill   = PatternFill("solid", fgColor=COLOR_POS)
    neg_fill   = PatternFill("solid", fgColor=COLOR_NEG)

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        for col in range(1, len(col_names) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = alt_fill()

        dev_val = ws.cell(row=i, column=5).value
        if dev_val is not None:
            ws.cell(row=i, column=5).number_format = '0.00"%"'
            if abs(dev_val) > 2.0:
                ws.cell(row=i, column=5).fill = alert_fill
            elif dev_val >= 0:
                ws.cell(row=i, column=5).fill = pos_fill
            else:
                ws.cell(row=i, column=5).fill = neg_fill

    autofit_columns(ws)


def write_monthly_returns_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Monthly Returns")

    if df.empty:
        ws["A1"] = "Brak danych"
        return

    # Pivot: wiersze = year_month, kolumny = symbol
    pivot = df.pivot_table(
        index="year_month", columns="symbol",
        values="monthly_return_pct", aggfunc="first"
    )
    pivot.reset_index(inplace=True)
    pivot.columns.name = None

    ws.append(list(pivot.columns))
    style_header_row(ws, 1, len(pivot.columns))
    ws.freeze_panes = "B2"

    pos_fill = PatternFill("solid", fgColor=COLOR_POS)
    neg_fill = PatternFill("solid", fgColor=COLOR_NEG)

    for i, row in enumerate(dataframe_to_rows(pivot, index=False, header=False), start=2):
        ws.append(row)
        for col in range(1, len(pivot.columns) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if col > 1 and cell.value is not None:
                cell.number_format = '0.00"%"'
                cell.fill = pos_fill if cell.value >= 0 else neg_fill
            if i % 2 == 0 and col == 1:
                cell.fill = alt_fill()

    autofit_columns(ws)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    logger.info("=== START: export_excel.py ===")

    wb = Workbook()
    wb.remove(wb.active)  # usuń domyślny pusty arkusz

    df_nav = get_nav_data()
    write_nav_sheet(wb, df_nav)
    logger.info(f"Arkusz 'Summary NAV': {len(df_nav)} tickerów")

    df_fx = get_fx_data()
    write_fx_sheet(wb, df_fx)
    logger.info(f"Arkusz 'Exchange Rates': {len(df_fx)} wierszy")

    df_monthly = get_monthly_returns()
    write_monthly_returns_sheet(wb, df_monthly)
    logger.info(f"Arkusz 'Monthly Returns': {len(df_monthly)} wierszy")

    wb.save(OUTPUT_FILE)
    logger.info(f"Zapisano raport: {OUTPUT_FILE}")
    logger.info("=== KONIEC: export_excel.py ===")


if __name__ == "__main__":
    main()
